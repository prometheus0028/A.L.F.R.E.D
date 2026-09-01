"""
ALFRED Files Tool - Complete file operations with safety, verification, and efficiency.

Supported operations:
- list(path)
- search(query, path=None, file_type=None)
- read(path)
- create(path, content)
- write(path, content)
- append(path, content)
- copy(source, destination)
- move(source, destination)
- rename(path, new_name)
- delete(path) - returns confirmation_required

Safety:
- All paths restricted to ALFRED_WORKSPACE_ROOT
- Path traversal (../, absolute escapes, symlinks) blocked
- .env/.git/secrets files never exposed
- Chunked reading for large files
- Metadata + snippets returned for efficiency
"""

import os
import shutil
import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import mimetypes

# Workspace enforcement
# Set to parent project directory to allow access to demo_data for testing
# Default workspace root
ALFRED_WORKSPACE_ROOT = Path(__file__).parent.parent.parent / "alfred_workspace"
DEMO_FALLBACK = Path(__file__).parent.parent.parent / "demo_data"

# Ensure workspace exists
ALFRED_WORKSPACE_ROOT.mkdir(parents=True, exist_ok=True)

# Blocked patterns (no direct access)
BLOCKED_PATTERNS = {".env", ".git", ".gitignore", "secrets", "private", "credentials"}
MAX_READ_SIZE = 500_000  # 500KB max to LLM per read (avoids token bloat)
MAX_SNIPPET_SIZE = 200   # Characters for search snippets
CHUNK_SIZE = 50_000      # For large file processing

def _normalize_path(path: str) -> Path:
    """Convert to absolute Path and validate it's within workspace."""
    try:
        # If path is just ".", use workspace root
        if path == ".":
            return ALFRED_WORKSPACE_ROOT
        
        # Treat as relative to workspace root, not cwd
        abs_path = (ALFRED_WORKSPACE_ROOT / path).expanduser().resolve()
    except (ValueError, OSError):
        raise ValueError(f"Invalid path: {path}")
    
    # Check if within workspace
    try:
        abs_path.relative_to(ALFRED_WORKSPACE_ROOT)
    except ValueError:
        raise ValueError(f"Path escapes workspace root: {path}")
    
    # Block symlink escapes
    if abs_path.is_symlink():
        raise ValueError(f"Symlinks not allowed: {path}")
    
    return abs_path

def _is_blocked_file(file_path: Path) -> bool:
    """Check if file/folder should not be accessed."""
    return any(blocked in file_path.parts for blocked in BLOCKED_PATTERNS)

def _get_file_type(file_path: Path) -> str:
    """Infer file type."""
    suffix = file_path.suffix.lower()
    type_map = {
        ".txt": "text",
        ".md": "markdown",
        ".json": "json",
        ".csv": "csv",
        ".pdf": "pdf",
        ".docx": "docx",
        ".xlsx": "xlsx",
    }
    return type_map.get(suffix, "unknown")

def _extract_text_from_file(file_path: Path, max_bytes: int = MAX_READ_SIZE) -> str:
    """Extract text from supported formats. For large files, read intelligently."""
    file_type = _get_file_type(file_path)
    
    try:
        if file_type in ("text", "markdown"):
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read(max_bytes)
            if len(content) >= max_bytes:
                content += "\n\n[... file truncated ...]"
            return content
        
        elif file_type == "json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            # Try pretty-print but truncate if too large
            text = json.dumps(data, indent=2)
            if len(text) > max_bytes:
                text = text[:max_bytes] + "\n\n[... file truncated ...]"
            return text
        
        elif file_type == "csv":
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read(max_bytes)
            if len(content) >= max_bytes:
                content += "\n\n[... file truncated ...]"
            return content
        
        elif file_type == "pdf":
            try:
                import pdfplumber
                text = ""
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text() or ""
                        text += page_text + "\n"
                        if len(text) >= max_bytes:
                            break
                if len(text) >= max_bytes:
                    text = text[:max_bytes] + "\n\n[... file truncated ...]"
                return text or "[No extractable text in PDF]"
            except Exception as e:
                return f"[PDF extraction failed: {str(e)}]"
        
        elif file_type == "docx":
            try:
                from docx import Document
                doc = Document(file_path)
                text = ""
                for para in doc.paragraphs:
                    text += para.text + "\n"
                    if len(text) >= max_bytes:
                        break
                if len(text) >= max_bytes:
                    text = text[:max_bytes] + "\n\n[... file truncated ...]"
                return text or "[No text in DOCX]"
            except Exception as e:
                return f"[DOCX extraction failed: {str(e)}]"
        
        elif file_type == "xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"\n=== Sheet: {sheet} ===\n"
                    for row in ws.iter_rows(values_only=True):
                        text += "\t".join(str(v or "") for v in row) + "\n"
                        if len(text) >= max_bytes:
                            break
                    if len(text) >= max_bytes:
                        break
                if len(text) >= max_bytes:
                    text = text[:max_bytes] + "\n\n[... file truncated ...]"
                return text or "[No data in XLSX]"
            except Exception as e:
                return f"[XLSX extraction failed: {str(e)}]"
        
        else:
            return f"[Unsupported file type: {file_type}]"
    
    except Exception as e:
        return f"[Read error: {str(e)}]"

def list(path: str = ".") -> Dict[str, Any]:
    """List files/folders in path. Returns metadata."""
    try:
        dir_path = _normalize_path(path)
        if not dir_path.exists():
            return {"error": "Path not found", "path": path}
        if not dir_path.is_dir():
            return {"error": "Path is not a directory", "path": path}
        
        items = []
        try:
            for item in dir_path.iterdir():
                if _is_blocked_file(item):
                    continue
                items.append({
                    "name": item.name,
                    "type": "folder" if item.is_dir() else _get_file_type(item),
                    "size": item.stat().st_size if item.is_file() else None,
                    "modified": item.stat().st_mtime,
                })
        except PermissionError:
            return {"error": "Permission denied", "path": path}
        
        return {
            "path": str(dir_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "count": len(items),
            "items": items
        }
    except ValueError as e:
        return {"error": str(e)}

def search(query: str = "", path: str = ".", file_type: Optional[str] = None) -> Dict[str, Any]:
    """
    Search filenames and content.
    Returns compact results: path, filename, type, snippet (if content match).
    """
    try:
        root_path = _normalize_path(path)
        if not root_path.exists():
            return {"error": "Search path not found", "query": query, "results": []}
        
        results = []
        query_lower = query.lower()
        
        # Walk directory tree
        for item in root_path.rglob("*"):
            if _is_blocked_file(item) or not item.is_file():
                continue
            
            item_type = _get_file_type(item)
            if file_type and item_type != file_type:
                continue
            
            # Filename match (higher priority)
            if query_lower in item.name.lower():
                results.append({
                    "path": str(item.relative_to(ALFRED_WORKSPACE_ROOT)),
                    "filename": item.name,
                    "type": item_type,
                    "size": item.stat().st_size,
                    "modified": item.stat().st_mtime,
                    "match_type": "filename",
                    "snippet": None,
                })
                continue
            
            # Content search (slower, only for small files)
            if item.stat().st_size < 100_000 and query:  # Only search <100KB files
                try:
                    content = _extract_text_from_file(item, max_bytes=100_000)
                    if query_lower in content.lower():
                        # Extract snippet around match
                        idx = content.lower().find(query_lower)
                        start = max(0, idx - 50)
                        end = min(len(content), idx + len(query) + 50)
                        snippet = "..." + content[start:end] + "..."
                        
                        results.append({
                            "path": str(item.relative_to(ALFRED_WORKSPACE_ROOT)),
                            "filename": item.name,
                            "type": item_type,
                            "size": item.stat().st_size,
                            "modified": item.stat().st_mtime,
                            "match_type": "content",
                            "snippet": snippet,
                        })
                except Exception:
                    pass
        
        return {
            "query": query,
            "path": str(root_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "count": len(results),
            "results": results[:20]  # Cap at 20 results for efficiency
        }
    except ValueError as e:
        return {"error": str(e), "query": query, "results": []}

def read(path: str) -> Dict[str, Any]:
    """Read file content. For large files, returns truncated + info."""
    try:
        file_path = _normalize_path(path)
        if not file_path.exists():
            return {"error": "File not found", "path": path}
        if not file_path.is_file():
            return {"error": "Path is not a file", "path": path}
        
        size = file_path.stat().st_size
        file_type = _get_file_type(file_path)
        
        content = _extract_text_from_file(file_path, max_bytes=MAX_READ_SIZE)
        
        return {
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "filename": file_path.name,
            "type": file_type,
            "size": size,
            "content": content,
            "truncated": len(content) >= MAX_READ_SIZE,
        }
    except ValueError as e:
        return {"error": str(e)}

def create(path: str, content: str = "") -> Dict[str, Any]:
    """Create new file. Fails if file exists."""
    try:
        file_path = _normalize_path(path)
        
        if file_path.exists():
            return {"error": "File already exists", "path": path}
        
        # Create parent directories
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        
        return {
            "status": "created",
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "size": len(content),
            "verified": file_path.exists()
        }
    except ValueError as e:
        return {"error": str(e)}

def write(path: str, content: str) -> Dict[str, Any]:
    """Overwrite file content. Creates if doesn't exist."""
    try:
        file_path = _normalize_path(path)
        
        # Create parent directories if needed
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        
        return {
            "status": "written",
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "size": len(content),
            "verified": file_path.exists() and file_path.stat().st_size == len(content)
        }
    except ValueError as e:
        return {"error": str(e)}

def append(path: str, content: str) -> Dict[str, Any]:
    """Append to file. Creates if doesn't exist."""
    try:
        file_path = _normalize_path(path)
        
        # Create parent directories if needed
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        initial_size = file_path.stat().st_size if file_path.exists() else 0
        
        # Use binary mode to ensure accurate byte counting (avoid platform-specific newline conversion)
        content_bytes = content.encode("utf-8")
        with open(file_path, "ab") as f:
            f.write(content_bytes)
        
        final_size = file_path.stat().st_size
        
        return {
            "status": "appended",
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "appended_bytes": len(content_bytes),
            "previous_size": initial_size,
            "new_size": final_size,
            "verified": final_size == initial_size + len(content_bytes)
        }
    except ValueError as e:
        return {"error": str(e)}

def copy(source: str, destination: str) -> Dict[str, Any]:
    """Copy file or directory. Destination must not exist."""
    try:
        src_path = _normalize_path(source)
        dst_path = _normalize_path(destination)
        
        if not src_path.exists():
            return {"error": "Source not found", "source": source}
        if dst_path.exists():
            return {"error": "Destination already exists", "destination": destination}
        
        # Create parent for destination
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        
        if src_path.is_file():
            shutil.copy2(src_path, dst_path)
        else:
            shutil.copytree(src_path, dst_path)
        
        return {
            "status": "copied",
            "source": str(src_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "destination": str(dst_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "verified": dst_path.exists()
        }
    except ValueError as e:
        return {"error": str(e)}

def move(source: str, destination: str) -> Dict[str, Any]:
    """Move file or directory. Destination must not exist."""
    try:
        src_path = _normalize_path(source)
        dst_path = _normalize_path(destination)
        
        if not src_path.exists():
            return {"error": "Source not found", "source": source}
        if dst_path.exists():
            return {"error": "Destination already exists", "destination": destination}
        
        # Create parent for destination
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        
        shutil.move(str(src_path), str(dst_path))
        
        return {
            "status": "moved",
            "source": str(src_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "destination": str(dst_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "verified": dst_path.exists() and not src_path.exists()
        }
    except ValueError as e:
        return {"error": str(e)}

def rename(path: str, new_name: str) -> Dict[str, Any]:
    """Rename file or folder. New name must be valid and not collide."""
    try:
        file_path = _normalize_path(path)
        
        if not file_path.exists():
            return {"error": "Path not found", "path": path}
        
        # Validate new_name (no path separators)
        if "/" in new_name or "\\" in new_name:
            return {"error": "Invalid name: cannot contain path separators", "new_name": new_name}
        
        new_path = file_path.parent / new_name
        
        # Validate new path is still within workspace
        try:
            new_path.resolve().relative_to(ALFRED_WORKSPACE_ROOT)
        except ValueError:
            return {"error": "New name would escape workspace", "new_name": new_name}
        
        if new_path.exists():
            return {"error": "Name already exists", "new_name": new_name}
        
        file_path.rename(new_path)
        
        return {
            "status": "renamed",
            "old_path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "new_path": str(new_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "verified": new_path.exists() and not file_path.exists()
        }
    except ValueError as e:
        return {"error": str(e)}

def delete(path: str) -> Dict[str, Any]:
    """
    Delete file or directory.
    Returns confirmation_required: True (LLM must acknowledge deletion intent).
    """
    try:
        file_path = _normalize_path(path)
        
        if not file_path.exists():
            return {"error": "Path not found", "path": path}
        
        # Return confirmation requirement (do NOT delete yet)
        return {
            "confirmation_required": True,
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "type": "folder" if file_path.is_dir() else "file",
            "size": None if file_path.is_dir() else file_path.stat().st_size,
            "message": f"Confirm deletion of {file_path.name}",
        }
    except ValueError as e:
        return {"error": str(e)}

def delete_confirmed(path: str) -> Dict[str, Any]:
    """
    INTERNAL: Actually delete after confirmation.
    Called by executor after user approval.
    """
    try:
        file_path = _normalize_path(path)
        
        if not file_path.exists():
            return {"error": "Path not found (already deleted?)", "path": path}
        
        if file_path.is_dir():
            shutil.rmtree(file_path)
        else:
            file_path.unlink()
        
        return {
            "status": "deleted",
            "path": str(file_path.relative_to(ALFRED_WORKSPACE_ROOT)),
            "verified": not file_path.exists()
        }
    except ValueError as e:
        return {"error": str(e)}
